from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT / "backend" / "data" / "Fichas_tecnicas-2026_04_11-23_57.xlsx"
OUTPUT_PATH = ROOT / "backend" / "data" / "inventory.json"

KNOWN_BRANDS = {
    "ngk": "NGK",
    "bosch": "Bosch",
    "denso": "Denso",
    "champion": "Champion",
    "autolite": "Autolite",
    "beru": "BERU",
    "wellman": "Wellman",
    "acdelco": "ACDelco",
}

PLUG_INCLUDE_RE = re.compile(
    r"buj|precalent|glow|spark|iridium|platin|v-power|laser",
    re.IGNORECASE,
)
PLUG_EXCLUDE_RE = re.compile(
    r"sensor|amort|correa|filtro|bomba|fusible|terminal|rodamiento|valvula|inyector|espejo|faro",
    re.IGNORECASE,
)

SHEETS = {
    "Bujias de encendido para veh...": "spark_plug",
    "Bujias de precalentamiento p...": "diesel_glow_plug",
    "Repuestos de linea liviana": "mixed",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text



def digits_only(value: Any) -> str:
    digits = re.sub(r"\D", "", normalize_text(value))
    return digits



def stable_seed(*parts: str) -> int:
    text = "|".join(parts)
    return sum((index + 1) * ord(char) for index, char in enumerate(text))



def infer_brand(raw_brand: str, title: str) -> str:
    candidate = normalize_text(raw_brand)
    if candidate and not candidate.isdigit() and "escribe" not in candidate.lower():
        lowered = candidate.lower()
        for key, canonical in KNOWN_BRANDS.items():
            if key in lowered:
                return canonical
        if len(candidate) <= 5:
            return candidate.upper()
        return candidate.title()

    lowered_title = title.lower()
    for key, canonical in KNOWN_BRANDS.items():
        if key in lowered_title:
            return canonical

    return "Generic"



def normalize_upc(*candidates: Any) -> str:
    for candidate in candidates:
        digits = digits_only(candidate)
        if 8 <= len(digits) <= 14:
            return digits
    return ""



def infer_vehicle_type(raw_value: Any) -> str:
    value = normalize_text(raw_value)
    if not value:
        return "Carro/Camioneta"
    return value



def infer_type(sheet_type: str, title: str) -> str | None:
    if sheet_type == "spark_plug":
        return "spark_plug"
    if sheet_type == "diesel_glow_plug":
        return "diesel_glow_plug"

    lowered = title.lower()
    if "precalent" in lowered or "glow" in lowered:
        return "diesel_glow_plug"
    if PLUG_INCLUDE_RE.search(lowered) and not PLUG_EXCLUDE_RE.search(lowered):
        return "spark_plug"
    return None



def infer_thread(item_type: str, row: dict[str, Any], vehicle_type: str) -> str:
    if item_type == "diesel_glow_plug":
        body_diameter = normalize_number(row.get("BODY_THREAD_DIAMETER"))
        unit = normalize_text(row.get("BODY_THREAD_DIAMETER_UNIT"))
        mm = convert_to_mm(body_diameter, unit)
        if mm and mm > 30 and body_diameter:
            mm = body_diameter
        if mm:
            return f"M{int(round(mm))} x 1.0"
        return "M10 x 1.0"

    lowered_vehicle = vehicle_type.lower()
    if "moto" in lowered_vehicle or "cuatriciclo" in lowered_vehicle:
        return "10mm x 1.0"
    return "14mm x 1.25"



def infer_reach(item_type: str, row: dict[str, Any], vehicle_type: str) -> str:
    if item_type == "diesel_glow_plug":
        length = normalize_number(row.get("GLOW_PLUG_LENGTH"))
        unit = normalize_text(row.get("GLOW_PLUG_LENGTH_UNIT"))
        mm = convert_to_mm(length, unit)
        if mm and mm > 400 and length:
            mm = length
        if mm:
            rounded = int(round(mm)) if abs(mm - round(mm)) < 0.05 else round(mm, 1)
            return f"{rounded}mm"
        return "65mm"

    lowered_vehicle = vehicle_type.lower()
    if "naut" in lowered_vehicle:
        return "12.7mm"
    return "19mm"



def infer_hex(item_type: str, vehicle_type: str) -> str:
    if item_type == "diesel_glow_plug":
        return "12mm"
    lowered_vehicle = vehicle_type.lower()
    if "naut" in lowered_vehicle:
        return "20.8mm"
    return "16mm"



def infer_gap(title: str, line: str, vehicle_type: str) -> float:
    lowered = f"{title} {line} {vehicle_type}".lower()
    if "iridium" in lowered or "laser" in lowered:
        return 0.8
    if "moto" in lowered:
        return 0.7
    if "v-power" in lowered or "copper" in lowered or "cooper" in lowered:
        return 1.1
    return 1.0



def infer_electrode(title: str, line: str) -> str:
    lowered = f"{title} {line}".lower()
    if "laser" in lowered and "iridium" in lowered:
        return "Laser Iridium"
    if "iridium" in lowered:
        return "Iridium"
    if "platin" in lowered:
        return "Platinum"
    if "v-power" in lowered:
        return "V-Power"
    if "copper" in lowered or "cooper" in lowered:
        return "Copper"
    if line:
        return normalize_text(line)
    return "Standard"



def infer_voltage(row: dict[str, Any]) -> str:
    value = normalize_number(row.get("GLOW_PLUG_VOLTAGE"))
    unit = normalize_text(row.get("GLOW_PLUG_VOLTAGE_UNIT")) or "V"
    if value:
        rounded = int(round(value)) if abs(value - round(value)) < 0.05 else round(value, 1)
        return f"{rounded}{unit}"
    return "11V"



def normalize_number(value: Any) -> float | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None



def convert_to_mm(value: float | None, unit: str) -> float | None:
    if value is None:
        return None
    lowered = unit.lower()
    if lowered == "cm":
        return value * 10
    return value



def infer_title(title: str, brand: str, model: str) -> str:
    if title:
        return title
    return f"{brand} {model}".strip()



def build_application(title: str, vehicle_type: str, origin: str) -> list[str]:
    applications: list[str] = []
    if title:
        applications.append(title)
        match = re.search(r"\bpara\b\s+(.+)$", title, re.IGNORECASE)
        if match:
            applications.append(match.group(1).strip(" .,-"))
    if vehicle_type:
        applications.append(vehicle_type)
    if origin:
        applications.append(origin)

    deduped: list[str] = []
    seen: set[str] = set()
    for value in applications:
        cleaned = normalize_text(value)
        key = cleaned.lower()
        if cleaned and key not in seen:
            deduped.append(cleaned)
            seen.add(key)
    return deduped or ["Aplicación no especificada"]



def simulate_price(item_type: str, brand: str, title: str, extra: str, quantity: int) -> float:
    lowered = f"{brand} {title} {extra}".lower()
    if item_type == "diesel_glow_plug":
        base = 16.5
        if "24v" in lowered:
            base = 23.5
        elif any(token in lowered for token in ("ngk", "beru", "bosch")):
            base = 19.5
    else:
        base = 4.4
        if "laser" in lowered and "iridium" in lowered:
            base = 11.4
        elif "iridium" in lowered:
            base = 9.8
        elif "platin" in lowered:
            base = 7.2
        elif "v-power" in lowered:
            base = 4.8
        elif "naut" in lowered:
            base = 5.3
        elif "moto" in lowered:
            base = 4.6

    if quantity > 1:
        base *= min(max(1, quantity), 8) * 0.78

    variation = (stable_seed(brand, title, extra) % 12) * 0.17
    return round(base + variation, 2)



def simulate_stock(sku: str, item_type: str) -> int:
    seed = stable_seed(sku, item_type)
    if item_type == "diesel_glow_plug":
        return 12 + (seed % 55)
    return 20 + (seed % 140)



def build_sku(brand: str, raw_sku: str, fallback: str) -> str:
    core = normalize_text(raw_sku) or normalize_text(fallback)
    core = re.sub(r"[^A-Za-z0-9.-]+", "-", core).strip("-").upper()
    brand_prefix = re.sub(r"[^A-Za-z0-9.-]+", "-", brand).strip("-").upper()
    if not core:
        return brand_prefix or "GENERIC"
    if core.startswith(brand_prefix):
        return core
    return f"{brand_prefix}-{core}" if brand_prefix else core



def build_record(row: dict[str, Any], sheet_name: str, sheet_type: str) -> dict[str, Any] | None:
    title = normalize_text(row.get("TITLE"))
    sku_raw = normalize_text(row.get("SKU"))
    part_number = normalize_text(row.get("PART_NUMBER")) or normalize_text(row.get("MODEL"))
    if not title and not sku_raw and not part_number:
        return None

    item_type = infer_type(sheet_type, title)
    if not item_type:
        return None

    brand = infer_brand(normalize_text(row.get("BRAND")), title)
    model = normalize_text(row.get("MODEL")) or part_number or sku_raw or "GENERIC"
    if model.lower() in {"iridium", "laser iridium", "v-power", "copper", "cooper", "standard"}:
        model = part_number or sku_raw or model
    vehicle_type = infer_vehicle_type(row.get("VEHICLE_TYPE"))
    origin = normalize_text(row.get("ORIGIN"))
    sku = build_sku(brand, sku_raw or model, part_number)
    if sku == re.sub(r"[^A-Za-z0-9.-]+", "-", brand).strip("-").upper() and model:
        sku = build_sku(brand, model, part_number)
    upc = normalize_upc(row.get("GTIN"), row.get("OEM"))
    title_clean = infer_title(title, brand, model)
    applications = build_application(title_clean, vehicle_type, origin)
    quantity = int(normalize_number(row.get("NUMBER_OF_SPARK_PLUGS_BY_KIT")) or 1)
    line = normalize_text(row.get("LINE"))

    record: dict[str, Any] = {
        "sku": sku,
        "upc": upc,
        "brand": brand,
        "model": model.upper(),
        "title": title_clean,
        "type": item_type,
        "thread": infer_thread(item_type, row, vehicle_type),
        "reach": infer_reach(item_type, row, vehicle_type),
        "hex": infer_hex(item_type, vehicle_type),
        "application": applications,
        "price_usd": simulate_price(item_type, brand, title_clean, line or vehicle_type, quantity),
        "stock": simulate_stock(sku, item_type),
        "vehicle_type": vehicle_type,
        "origin": origin,
        "source_sheet": sheet_name,
        "source_publication_id": normalize_text(row.get("ID")),
    }

    if item_type == "spark_plug":
        record["gap_mm"] = infer_gap(title_clean, line, vehicle_type)
        record["electrode"] = infer_electrode(title_clean, line)
        if quantity > 1:
            record["kit_quantity"] = quantity
    else:
        record["voltage"] = infer_voltage(row)
        record["electrode"] = normalize_text(row.get("GLOW_PLUG_MATERIAL")) or "Cobre"

    return record



def load_rows() -> list[dict[str, Any]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    records_by_sku: dict[str, dict[str, Any]] = {}

    for sheet_name, sheet_type in SHEETS.items():
        worksheet = workbook[sheet_name]
        headers = [normalize_text(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]

        for raw_row in worksheet.iter_rows(min_row=5, values_only=True):
            row = {headers[index]: raw_row[index] if index < len(raw_row) else None for index in range(len(headers))}
            record = build_record(row, sheet_name, sheet_type)
            if not record:
                continue

            existing = records_by_sku.get(record["sku"])
            if existing:
                existing_apps = set(existing.get("application", []))
                existing["application"] = existing.get("application", []) + [
                    app for app in record["application"] if app not in existing_apps
                ]
                if not existing.get("upc") and record.get("upc"):
                    existing["upc"] = record["upc"]
                continue

            records_by_sku[record["sku"]] = record

    records = sorted(records_by_sku.values(), key=lambda item: (item["brand"], item["model"], item["sku"]))
    return records



def main() -> None:
    records = load_rows()
    OUTPUT_PATH.write_text(json.dumps({"items": records}, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    spark_count = sum(1 for item in records if item["type"] == "spark_plug")
    glow_count = sum(1 for item in records if item["type"] == "diesel_glow_plug")
    print(f"Wrote {len(records)} items to {OUTPUT_PATH}")
    print(f"Spark plugs: {spark_count}")
    print(f"Glow plugs: {glow_count}")


if __name__ == "__main__":
    main()
